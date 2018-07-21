from math import sin, cos, sqrt, atan2, radians
import pandas as pd

# approximate radius of earth in meters
def distance_cal(LatC,LongC,LatT,LongT):
    R = 6373.0

    lat1 = radians(LatC)
    lon1 = radians(LongC)
    lat2 = radians(LatT)
    lon2 = radians(LongT)

    dlon = lon2 - lon1
    dlat = lat2 - lat1

    a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    distance = R * c*1000

#    print("Result:", distance)
    return distance

#Read the input files
condo=pd.read_excel('Condo_info_Combined - Cleaned.xlsx')
train=pd.read_excel('TrainsLatLong.xlsx')

CondoName,LatC,LongC=condo['Name'],condo['Latitude'],condo['Longitude']    
StationName,LatT,LongT=train['StationName'],train['LatT'] ,train['LongT'] 

print(len(CondoName))
print(len(set(CondoName)))
print(len(StationName))

#Loop to check distance to stations for all condos and return only the closest station

c=0
NearestStation=[]

while c <= len(CondoName):
    station=""
    dist=9999999
    
    for i in range(0,len(StationName)): #loop check distance for this condo
        result=distance_cal(LatC[c],LongC[c],LatT[i],LongT[i])
        if result < dist:
            dist= result
            station= StationName[i]
            print(dist)
    print(str(c)+" The Nearest Train Station for "+CondoName[c]+ " is "+station,str(dist)+" m.")
    NearestStation.append([station,dist])
    c+=1
    
    try:
        while CondoName[c]==CondoName[c-1]: # if the next row is the same condo, copy over information. Exit loop condo name mismatch.
            NearestStation.append([station,dist])
            print(str(c)+" The Nearest Train Station for "+CondoName[c]+ " is "+station,str(dist)+" m.")
            c+=1
        print('============= Next Condo')
        print('c = '+str(c))
    except: break
print(NearestStation)
    

# save result to excel file
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Condo_info"
dest_filename = 'Added_dist.xlsx'

for r in range(0,len(NearestStation)):
    ws.cell(row=r+2,column=1).value = NearestStation[r][0]
    ws.cell(row=r+2,column=2).value = NearestStation[r][1]

wb.save(filename = dest_filename)


